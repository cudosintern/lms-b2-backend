from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from app.core.database import get_db
from .cons_absentees_report_schema import *
from sqlalchemy import text, bindparam
from fastapi.responses import FileResponse,StreamingResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from datetime import datetime
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.styles import Font
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
import os

router = APIRouter(tags=["Consolidated Absentees Report"])

# --------------------------------------------------
# 1. DROPDOWNS
# --------------------------------------------------

@router.get("/departments")
def get_departments(db: Session = Depends(get_db)):
    try:
        result = db.execute(text("""
            SELECT 
                dept_id AS id,
                dept_name AS name
            FROM iems_department
            WHERE status = 1
        """)).fetchall()

        return [dict(row._mapping) for row in result]

    except Exception as e:
        print("ERROR:", str(e))
        return {"error": str(e)}
    

@router.get("/programs/{department_id}")
def get_programs(department_id: int, db: Session = Depends(get_db)):
    try:
        result = db.execute(text("""
            SELECT 
                pgm_id AS id,
                pgm_title AS name
            FROM iems_program
            WHERE dept_id = :dept
            AND status = 1
        """), {"dept": department_id}).fetchall()

        return [dict(row._mapping) for row in result]

    except Exception as e:
        print("ERROR:", str(e))
        return {"error": str(e)}

@router.get("/curriculum/{program_id}")
def get_curriculum(program_id: int, db: Session = Depends(get_db)):
    try:
        result = db.execute(text("""
            SELECT 
                academic_batch_id AS id,
                academic_batch_desc AS name
            FROM iems_academic_batch
            WHERE pgm_id = :prog
            AND status = 1
        """), {"prog": program_id}).fetchall()

        return [dict(row._mapping) for row in result]

    except Exception as e:
        print("ERROR:", str(e))
        return {"error": str(e)}

@router.get("/terms/{curriculum_id}")
def get_terms(curriculum_id: int, db: Session = Depends(get_db)):
    result = db.execute(text("""
        SELECT 
            semester_id AS id,
            semester_desc AS name
        FROM iems_semester
        ORDER BY semester
    """)).fetchall()

    return [dict(row._mapping) for row in result]

@router.get("/sections/{semester_id}")
def get_sections(semester_id: int, db: Session = Depends(get_db)):
    result = db.execute(text("""
        SELECT id, section AS name
        FROM iems_section
        WHERE semester_id = :term
    """), {"term": semester_id}).fetchall()

    return [{"id": row[0], "name": row[1]} for row in result]
# --------------------------------------------------
# 2. DEFAULT DATE API
# --------------------------------------------------

@router.get("/date-info")
def get_date_info(db: Session = Depends(get_db)):

    latest_date = db.execute(text("""
        SELECT MAX(attendance_date) as latest_date
        FROM lms_manage_attendance
    """)).scalar()

    scheduled_dates = db.execute(text("""
        SELECT DISTINCT attendance_date
        FROM lms_manage_attendance
    """)).fetchall()

    return {
        "latest_attendance_date": latest_date,
        "scheduled_dates": [d[0] for d in scheduled_dates]
    }


# --------------------------------------------------
# 3. MAIN REPORT
# --------------------------------------------------

@router.post("/report")
def get_report(payload: ReportRequest, db: Session = Depends(get_db)):

    query = """
    SELECT
    d.dept_name AS department,
    s.term_name AS term,
    c.crs_title AS course,
    sec.section AS section,
    COUNT(sa.stud_attendance_id) AS absent_count,
    c.crs_id AS course_id,
    sec.id AS section_id

    FROM lms_manage_attendance ma

    JOIN lms_map_student_attendance sa
        ON sa.attendance_id = ma.attendance_id

    JOIN iems_section sec
        ON sec.id = ma.section_id

    JOIN iems_semester s
        ON s.semester_id = sec.semester_id

    JOIN iems_academic_batch ab
        ON ab.academic_batch_id = s.academic_batch_id

    JOIN iems_program p
        ON p.pgm_id = ab.pgm_id

    JOIN iems_department d
        ON d.dept_id = p.dept_id

    JOIN iems_students st
        ON st.usno = sa.student_usn

    JOIN iems_courses c
        ON c.crs_id = ma.crs_id

    WHERE sa.attendance_status = 'ABSENT'

    AND STR_TO_DATE(ma.attendance_date, '%Y-%m-%d')
        BETWEEN :start AND :end
    """

    params = {
        "start": payload.start_date,
        "end": payload.end_date
    }

    # ✅ Dynamic filters (ONLY here)
    if payload.department_ids is not None:
        query += " AND d.dept_id IN :dept_ids"
        params["dept_ids"] = payload.department_ids

    if payload.program_ids is not None:
        query += " AND p.pgm_id IN :prog"
        params["prog"] = payload.program_ids

    if payload.curriculum_ids is not None:
        query += " AND ab.academic_batch_id IN :cur"
        params["cur"] = payload.curriculum_ids

    if payload.semester_ids is not None:
        query += " AND s.semester_id IN :sem"
        params["sem"] = payload.semester_ids

    if payload.section_ids is not None:
        query += " AND sec.id IN :sec"
        params["sec"] = payload.section_ids

    # ✅ ONLY ONE GROUP BY (at the end)
    query += """
    GROUP BY d.dept_name, s.term_name, c.crs_title, sec.section
    """

    stmt = text(query)

    # ✅ Bind expanding params ONLY if they exist
    bind_params = []

    if payload.department_ids:
        bind_params.append(bindparam("dept_ids", expanding=True))

    if payload.program_ids:
        bind_params.append(bindparam("prog", expanding=True))

    if payload.curriculum_ids:
        bind_params.append(bindparam("cur", expanding=True))

    if payload.semester_ids:
        bind_params.append(bindparam("sem", expanding=True))

    if payload.section_ids:
        bind_params.append(bindparam("sec", expanding=True))

    if bind_params:
        stmt = stmt.bindparams(*bind_params)

    result = db.execute(stmt, params).fetchall()

    return [dict(row._mapping) for row in result]
# --------------------------------------------------
# 4. DRILLDOWN
# --------------------------------------------------

@router.post("/drilldown")
def get_drilldown(payload: DrilldownRequest, db: Session = Depends(get_db)):

    result = db.execute(text("""
   SELECT
    ma.attendance_date,
    st.name AS student_name,
    st.usno,
    st.mobile

FROM lms_manage_attendance ma

JOIN lms_map_student_attendance sa
    ON sa.attendance_id = ma.attendance_id

JOIN iems_students st
    ON st.usno = sa.student_usn   -- ✅ FIXED

WHERE sa.attendance_status = 'ABSENT'   -- ✅ FIXED

AND ma.crs_id = :course
AND ma.section_id = :section

AND STR_TO_DATE(ma.attendance_date, '%Y-%m-%d')
    BETWEEN :start AND :end

    ORDER BY ma.attendance_date
    """), {
        "course": payload.course_id,
        "section": payload.section_id,
        "start": payload.start_date,
        "end": payload.end_date
    }).fetchall()

    return [dict(row._mapping) for row in result]

@router.post("/export/xls")
def export_absentees_xls(data: ReportRequest, db: Session = Depends(get_db)):
    try:
        base_query = """
        SELECT
            d.dept_name AS department,
            s.term_name AS term,
            c.crs_title AS course,
            sec.section AS section,
            COUNT(sa.stud_attendance_id) AS absent_count

        FROM lms_manage_attendance ma
        JOIN lms_map_student_attendance sa ON sa.attendance_id = ma.attendance_id
        JOIN iems_section sec ON sec.id = ma.section_id
        JOIN iems_semester s ON s.semester_id = sec.semester_id
        JOIN iems_academic_batch ab ON ab.academic_batch_id = s.academic_batch_id
        JOIN iems_program p ON p.pgm_id = ab.pgm_id
        JOIN iems_department d ON d.dept_id = p.dept_id
        JOIN iems_courses c ON c.crs_id = ma.crs_id

        WHERE sa.attendance_status = 'ABSENT'
        AND STR_TO_DATE(ma.attendance_date, '%Y-%m-%d')
        BETWEEN :start AND :end
        """

        params = {
            "start": data.start_date,
            "end": data.end_date
        }

        # 🔥 DYNAMIC FILTERS
        if data.department_ids:
            base_query += f" AND d.dept_id IN ({','.join(map(str, data.department_ids))})"

        if data.program_ids:
            base_query += f" AND p.pgm_id IN ({','.join(map(str, data.program_ids))})"

        if data.curriculum_ids:
            base_query += f" AND ab.academic_batch_id IN ({','.join(map(str, data.curriculum_ids))})"

        if data.semester_ids:
            base_query += f" AND s.semester_id IN ({','.join(map(str, data.semester_ids))})"

        if data.section_ids:
            base_query += f" AND sec.id IN ({','.join(map(str, data.section_ids))})"

        base_query += " GROUP BY d.dept_name, s.term_name, c.crs_title, sec.section"

        result = db.execute(text(base_query), params).mappings().all()

        # ✅ SAFETY
        if not result:
            return Response(content=b"", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # ── Excel ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Absentees Report"

        ws["A1"] = "Consolidated Absentees Report"
        ws["A1"].font = Font(bold=True)

        ws.append(["Date", datetime.now().strftime("%d-%m-%Y")])
        ws.append([])

        headers = ["Department", "Term", "Course", "Section", "Absent Count"]
        ws.append(headers)

        for col in range(1, 6):
            ws.cell(row=4, column=col).font = Font(bold=True)

        for row in result:
            ws.append([
                row["department"],
                row["term"],
                row["course"],
                row["section"],
                row["absent_count"]
            ])

        file_path = f"./absentees.xlsx"
        wb.save(file_path)

        return FileResponse(file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Absentees_Report.xlsx"
        )

    except Exception as e:
        print("XLS ERROR:", e)
        return {"error": str(e)}



@router.post("/export/pdf")
def export_absentees_pdf(data: ReportRequest, db: Session = Depends(get_db)):
    try:
        base_query = """
        SELECT
            d.dept_name AS department,
            s.term_name AS term,
            c.crs_title AS course,
            sec.section AS section,
            COUNT(sa.stud_attendance_id) AS absent_count

        FROM lms_manage_attendance ma
        JOIN lms_map_student_attendance sa ON sa.attendance_id = ma.attendance_id
        JOIN iems_section sec ON sec.id = ma.section_id
        JOIN iems_semester s ON s.semester_id = sec.semester_id
        JOIN iems_academic_batch ab ON ab.academic_batch_id = s.academic_batch_id
        JOIN iems_program p ON p.pgm_id = ab.pgm_id
        JOIN iems_department d ON d.dept_id = p.dept_id
        JOIN iems_courses c ON c.crs_id = ma.crs_id

        WHERE sa.attendance_status = 'ABSENT'
        AND STR_TO_DATE(ma.attendance_date, '%Y-%m-%d')
        BETWEEN :start AND :end
        """

        params = {
            "start": data.start_date,
            "end": data.end_date
        }

        # 🔥 SAME FILTER LOGIC
        if data.department_ids:
            base_query += f" AND d.dept_id IN ({','.join(map(str, data.department_ids))})"

        if data.program_ids:
            base_query += f" AND p.pgm_id IN ({','.join(map(str, data.program_ids))})"

        if data.curriculum_ids:
            base_query += f" AND ab.academic_batch_id IN ({','.join(map(str, data.curriculum_ids))})"

        if data.semester_ids:
            base_query += f" AND s.semester_id IN ({','.join(map(str, data.semester_ids))})"

        if data.section_ids:
            base_query += f" AND sec.id IN ({','.join(map(str, data.section_ids))})"

        base_query += " GROUP BY d.dept_name, s.term_name, c.crs_title, sec.section"

        result = db.execute(text(base_query), params).mappings().all()

        # ✅ SAFETY
        if not result:
            return Response(content=b"", media_type="application/pdf")

        file_path = "./absentees.pdf"

        doc = SimpleDocTemplate(file_path)
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph("<b>Consolidated Absentees Report</b>", styles["Title"]))

        table_data = [["Department", "Term", "Course", "Section", "Absent Count"]]

        for row in result:
            table_data.append([
                row["department"],
                row["term"],
                row["course"],
                row["section"],
                row["absent_count"]
            ])

        table = Table(table_data)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))

        elements.append(table)
        doc.build(elements)

        return FileResponse(file_path, media_type="application/pdf", filename="Absentees_Report.pdf")

    except Exception as e:
        print("PDF ERROR:", e)
        return {"error": str(e)}